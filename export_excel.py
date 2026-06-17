"""
업로드 완료된 번호(uploaded=1)를 일자별로 엑셀로 내보낸다.
output/YYYY-MM-DD.xlsx 형식으로 저장
"""
import sqlite3
import os
from datetime import datetime
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

DB_PATH = os.path.join(os.path.dirname(__file__), "missed_calls.db")
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "output")


def normalize_phone(number: str) -> str:
    digits = "".join(filter(str.isdigit, number))
    if len(digits) == 11 and digits.startswith("010"):
        return f"{digits[:3]}-{digits[3:7]}-{digits[7:]}"
    return number


def export():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    conn = sqlite3.connect(DB_PATH)
    rows = conn.execute("""
        SELECT number, device_id, received_at, saved_at
        FROM calls
        WHERE uploaded = 1
        ORDER BY saved_at ASC
    """).fetchall()
    conn.close()

    if not rows:
        print("업로드 완료된 데이터가 없습니다.")
        return

    # 날짜별로 그룹화 (saved_at 기준)
    by_date = defaultdict(list)
    for number, device_id, received_at, saved_at in rows:
        date = saved_at[:10]  # YYYY-MM-DD
        by_date[date].append((number, device_id, received_at, saved_at))

    for date, items in sorted(by_date.items()):
        path = os.path.join(OUTPUT_DIR, f"{date}.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = date

        # 헤더
        headers = ["번호", "기기", "수신시각", "저장시각"]
        header_fill = PatternFill("solid", fgColor="4472C4")
        header_font = Font(bold=True, color="FFFFFF")
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # 데이터
        for row_idx, (number, device_id, received_at, saved_at) in enumerate(items, 2):
            ws.cell(row=row_idx, column=1, value=normalize_phone(number))
            ws.cell(row=row_idx, column=2, value=device_id)
            ws.cell(row=row_idx, column=3, value=received_at)
            ws.cell(row=row_idx, column=4, value=saved_at)

        # 열 너비 자동 조정
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 22

        wb.save(path)
        print(f"저장: {path} ({len(items)}건)")

    print(f"\n총 {len(by_date)}개 파일 생성 완료")


if __name__ == "__main__":
    export()
